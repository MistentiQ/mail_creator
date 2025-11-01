import os
from PySimpleGUI import Window, Input, popup, WIN_CLOSED, FileBrowse, Button, Column, InputText, theme, Text,one_line_progress_meter
import pyperclip
from dev import *
from dev.excel import write_excel
from dev.user_pass import generate_random_password
from openpyxl import load_workbook

theme('DarkGrey8')  
# Вкладки
layout1 = [[Text('Добавление нового пользователя')],
           [Text('Имя пользователя')],
           [InputText(key='Username')],
           [Text('Описание')],
           [InputText(key='Dis')],
           [Button('    Записать    ')],
           [Text('')]]

layout2 = [[Text('Добавление списка пользователей')],
           [Input(key='path'), FileBrowse()],
           [Button('Загрузить список')]]

layout3 = [[Text('Смена пароля')],
           [Text('В разработке')]]

# Основная страница
layout = [[Button('По одному'), Button('Списком'), Button('Смена пароля')],
          [Column(layout1, key='-COL1-'), Column(layout2, visible=False, key='-COL2-'),
           Column(layout3, visible=False, key='-COL3-')],
          [Button('Cancel'), Button('Показать готовые')]]

window = Window('Новые пользователи почты', layout)


def main():
    while True:
        event, values = window.read()
        # кнопки переключения вкладок
        if event == 'По одному':
            window[f'-COL2-'].update(visible=False)
            window[f'-COL1-'].update(visible=True)
            window[f'-COL3-'].update(visible=False)
        if event == 'Списком':
            window[f'-COL1-'].update(visible=False)
            window[f'-COL2-'].update(visible=True)
            window[f'-COL3-'].update(visible=False)
        if event == 'Смена пароля':
            window[f'-COL3-'].update(visible=True)
            window[f'-COL2-'].update(visible=False)
            window[f'-COL1-'].update(visible=False)
        # код основной вкладки
        if event == '    Записать    ':
            username = values['Username']
            discript = values['Dis']
            if username == "" or discript == "":
                popup("Error:", "Имя пользователя и Описание", "должны быть заполнены")
                main()
            userpassword = generate_random_password()
            email = str(values['Username']) + "@zdrav.spb.ru"
            # Пишем Excel
            write_excel(username, userpassword, email, discript)
            # Создаём док для выдачи
            create_word_file(username, userpassword, email, discript)
            # Заносим юзверя в базу
            adduser_mssql(username, userpassword, email, discript)
            # Говорим пользователю что у нас всё получилось
            pyperclip.copy("Пароль архива: " + PASSWORD)  # закидываем пароль в буфер обмена
            popup("Пароль пользователя: ", PASSWORD, '\n', "Пароль скопирован!")
            main()

        if event == 'Загрузить список':
            path = (values['path'])  # без этой переменной не хочет видеть путь
            wb2 = load_workbook(path)
            ws = wb2.active
            row_count = ws.max_row - 1
            stroka = int(2)
            try:
                for i in range(row_count):
                    one_line_progress_meter('Добавление пользователей', i+1, row_count, orientation='h', bar_color='gray on white')
                    cell = ws.cell(row=stroka, column=1)
                    cell2 = ws.cell(row=stroka, column=2)
                    username = cell.value
                    discript = cell2.value
                    email = str(cell.value) + "@zdrav.spb.ru"
                    userpassword = generate_random_password()
                    # Пишем Excel
                    write_excel(username, userpassword, email, discript)
                    # Создаём док для выдачи
                    create_word_file_list(username, userpassword, email, discript)
                    # Заносим юзверя в базу
                    adduser_mssql(username, userpassword, email, discript)
                    # Говорим пользователю что у нас всё получилось
                    stroka = int(stroka) + 1

            finally:
                popup(stroka - 2, 'Пользователей', 'добавлены в базу')

        if event == 'Показать готовые':
            os.system("explorer \\\\miacshare\\mailzdrav$\\Готовые")
            main()

        if event == WIN_CLOSED or event == 'Cancel':
            break


if __name__ == "__main__":
    main()
